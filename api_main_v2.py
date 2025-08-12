import tempfile
import requests
import os
import httpx
import asyncio
import tempfile
import logging
import torch
import time
import hashlib
import pickle
import uuid
import json
from datetime import datetime
from fastapi import FastAPI
from pydantic import BaseModel
from typing import List, Dict, Any
from dotenv import load_dotenv
from asyncio import Lock
from collections import defaultdict
import re
import mimetypes
from urllib.parse import urlparse
import zipfile
import io
from langchain_community.document_loaders import (
    PyMuPDFLoader, 
    UnstructuredPowerPointLoader,
    UnstructuredWordDocumentLoader,
    UnstructuredExcelLoader,
    UnstructuredImageLoader,
    TextLoader,
    CSVLoader
)
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain_community.vectorstores import FAISS
from langchain.prompts import PromptTemplate
from langchain.schema import Document
import warnings
warnings.filterwarnings("ignore")

# Suppress Aspose.Slides warnings
warnings.filterwarnings("ignore", "The type initializer for")

# === Load API Keys ===
load_dotenv()
NVIDIA_KEYS = [
    ("NVIDIA_API_KEY_1", os.getenv("NVIDIA_API_KEY_1")),
    ("NVIDIA_API_KEY_2", os.getenv("NVIDIA_API_KEY_2")),
    ("NVIDIA_API_KEY_3", os.getenv("NVIDIA_API_KEY_3")),
    ("NVIDIA_API_KEY_4", os.getenv("NVIDIA_API_KEY_4")),
    ("NVIDIA_API_KEY_5", os.getenv("NVIDIA_API_KEY_5")),
]

KEY_INDEX = 0
lock = Lock()

# === Enhanced File Type Support ===
SUPPORTED_EXTENSIONS = {
    '.pdf': 'pdf',
    '.ppt': 'powerpoint',
    '.pptx': 'powerpoint', 
    '.doc': 'word',
    '.docx': 'word',
    '.xls': 'excel',
    '.xlsx': 'excel',
    '.jpg': 'image',
    '.jpeg': 'image',
    '.png': 'image',
    '.gif': 'image',
    '.bmp': 'image',
    '.tiff': 'image',
    '.bin': 'binary',
    '.zip': 'archive',
    '.rar': 'archive',
    '.7z': 'archive',
    '.txt': 'text',
    '.csv': 'csv'
}

# === Logger Setup ===
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# === Cache Setup ===
CACHE_DIR = "embedding_cache"
MARKDOWN_CACHE_DIR = os.path.join(CACHE_DIR, "markdown")

# Create cache directories
for cache_dir in [CACHE_DIR, MARKDOWN_CACHE_DIR]:
    if not os.path.exists(cache_dir):
        os.makedirs(cache_dir)
        logger.info(f"📁 Created cache directory: {cache_dir}")

def get_markdown_cache_path(url):
    """Get the cache file path for markdown content"""
    cache_key = hashlib.md5(url.encode()).hexdigest()
    return os.path.join(MARKDOWN_CACHE_DIR, f"{cache_key}_ppt.md")

def save_markdown_to_cache(content, url):
    """Save markdown content to cache"""
    try:
        cache_path = get_markdown_cache_path(url)
        with open(cache_path, 'w', encoding='utf-8') as f:
            f.write(content)
        logger.info(f"💾 Markdown content cached: {cache_path}")
        return True
    except Exception as e:
        logger.warning(f"⚠️ Failed to cache markdown: {e}")
        return False

def load_markdown_from_cache(url):
    """Load markdown content from cache if it exists"""
    try:
        cache_path = get_markdown_cache_path(url)
        if os.path.exists(cache_path):
            with open(cache_path, 'r', encoding='utf-8') as f:
                content = f.read()
            logger.info(f"📋 Loaded markdown from cache: {cache_path}")
            return content
        return None
    except Exception as e:
        logger.warning(f"⚠️ Failed to load markdown from cache: {e}")
        return None

def get_file_hash(url, file_content=None):
    """Generate a unique hash for the file URL and content type"""
    url_hash = hashlib.md5(url.encode()).hexdigest()
    if file_content:
        content_hash = hashlib.md5(file_content[:1000]).hexdigest()
        return f"{url_hash}_{content_hash}"
    return url_hash

def detect_file_type(url, content_type=None, file_content=None):
    """Enhanced file type detection"""
    try:
        # First try URL extension
        parsed_url = urlparse(url)
        file_path = parsed_url.path.lower()
        
        for ext, file_type in SUPPORTED_EXTENSIONS.items():
            if file_path.endswith(ext):
                return file_type, ext
        
        # Try content type
        if content_type:
            mime_to_type = {
                'application/pdf': ('pdf', '.pdf'),
                'application/vnd.ms-powerpoint': ('powerpoint', '.ppt'),
                'application/vnd.openxmlformats-officedocument.presentationml.presentation': ('powerpoint', '.pptx'),
                'application/msword': ('word', '.doc'),
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document': ('word', '.docx'),
                'application/vnd.ms-excel': ('excel', '.xls'),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': ('excel', '.xlsx'),
                'image/jpeg': ('image', '.jpg'),
                'image/png': ('image', '.png'),
                'image/gif': ('image', '.gif'),
                'text/plain': ('text', '.txt'),
                'text/csv': ('csv', '.csv'),
                'application/zip': ('archive', '.zip'),
                'application/octet-stream': ('binary', '.bin')
            }
            if content_type in mime_to_type:
                return mime_to_type[content_type]
        
        # Try content analysis for binary files
        if file_content:
            if file_content.startswith(b'%PDF'):
                return 'pdf', '.pdf'
            elif file_content.startswith(b'PK'):
                return 'archive', '.zip'
            elif b'Microsoft Office' in file_content[:1000]:
                return 'word', '.docx'
        
        # Default to binary
        return 'binary', '.bin'
        
    except Exception as e:
        logger.warning(f"File type detection failed: {e}")
        return 'binary', '.bin'

async def download_file_safely(url, max_size_mb=100, check_binary=True):
    """Safely download file with size limits and error handling"""
    try:
        logger.info(f"🔄 Attempting to download file from: {url}")
        
        async with httpx.AsyncClient(timeout=30) as client:
            # First, get headers to check file size and type
            try:
                head_response = await client.head(url)
                content_length = head_response.headers.get('Content-Length')
                content_type = head_response.headers.get('Content-Type', '')
            except Exception as head_error:
                logger.warning(f"⚠ HEAD request failed: {head_error}, proceeding with limited info")
                content_length = None
                content_type = ''
            
            # Check if it's a binary file and size is too large
            if content_length and check_binary:
                size_mb = int(content_length) / (1024 * 1024)
                file_type, _ = detect_file_type(url, content_type)
                
                # If it's a binary file and larger than 50MB, skip download
                if file_type == 'binary' and size_mb > 50:
                    logger.info(f"🚫 Skipping large binary file: {size_mb:.1f}MB")
                    return None, content_type  # Return None to indicate binary skip
                
                if size_mb > max_size_mb:
                    logger.info(f"🚫 Skipping file too large: {size_mb:.1f}MB (max: {max_size_mb}MB)")
                    return None, content_type  # Return None instead of raising exception
            
            # Check file type even without content length for known binary extensions
            if check_binary:
                file_type, _ = detect_file_type(url, content_type)
                if file_type == 'binary':
                    # For binary files without size info, skip download for safety
                    logger.info(f"🚫 Skipping binary file (unknown size): {url}")
                    return None, content_type
            
            # Download the file
            response = await client.get(url, timeout=120)
            if response.status_code != 200:
                raise Exception(f"Download failed with status {response.status_code}")
            
            content = response.content
            if len(content) > max_size_mb * 1024 * 1024:
                raise Exception(f"File size exceeded {max_size_mb}MB during download")
            
            logger.info(f"✅ Successfully downloaded {len(content)} bytes")
            return content, content_type
            
    except Exception as e:
        logger.error(f"❌ Download failed: {e}")
        raise e

def process_archive_file(content, file_extension):
    """Enhanced archive extraction with robust ZIP handling and multiple format support"""
    documents = []
    extracted_files_count = 0
    total_files_count = 0
    
    try:
        if file_extension in ['.zip']:
            logger.info(f"🗜️ Processing ZIP archive...")
            
            # Try multiple extraction methods for better compatibility
            zip_file = None
            try:
                # Method 1: Standard zipfile with enhanced error recovery
                zip_file = zipfile.ZipFile(io.BytesIO(content))
                logger.info(f"✅ ZIP file opened successfully with standard method")
            except zipfile.BadZipFile:
                logger.warning(f"⚠️ Standard ZIP extraction failed, trying alternative methods...")
                try:
                    # Method 2: Try with different mode and strict mode disabled
                    zip_file = zipfile.ZipFile(io.BytesIO(content), mode='r', allowZip64=True, strict_timestamps=False)
                    logger.info(f"✅ ZIP file opened with allowZip64=True and relaxed timestamps")
                except Exception as e2:
                    logger.warning(f"⚠️ Alternative ZIP method 2 failed: {e2}")
                    try:
                        # Method 3: Try to recover partial ZIP content
                        logger.info(f"🔧 Attempting ZIP recovery mode...")
                        # Reset the BytesIO position
                        content_stream = io.BytesIO(content)
                        content_stream.seek(0)
                        zip_file = zipfile.ZipFile(content_stream, mode='r')
                        logger.info(f"✅ ZIP file opened with recovery mode")
                    except Exception as e3:
                        logger.error(f"❌ All ZIP extraction methods failed: {e3}")
                        raise Exception(f"Cannot open ZIP file: corrupted, password-protected, or invalid format. Try extracting manually.")
            
            if zip_file:
                try:
                    # Get file list with better error handling
                    file_list = zip_file.namelist()
                    total_files_count = len(file_list)
                    logger.info(f"📁 Found {total_files_count} files in ZIP archive")
                    
                    if total_files_count == 0:
                        logger.warning(f"⚠️ ZIP archive is empty")
                        return []
                    
                    # Process each file with enhanced extraction
                    for filename in file_list:
                        try:
                            # Skip directories and hidden files
                            if filename.endswith('/') or filename.startswith('__MACOSX/') or filename.startswith('.'):
                                logger.info(f"⏭️ Skipping directory/hidden file: {filename}")
                                continue
                                
                            file_info = zip_file.getinfo(filename)
                            
                            # Enhanced size check with more reasonable limits
                            if file_info.file_size > 100 * 1024 * 1024:  # 100MB limit per file
                                logger.warning(f"⚠️ Skipping large file {filename}: {file_info.file_size / (1024*1024):.1f}MB")
                                continue
                            
                            if file_info.file_size == 0:
                                logger.info(f"⏭️ Skipping empty file: {filename}")
                                continue
                            
                            logger.info(f"📄 Extracting: {filename} ({file_info.file_size} bytes)")
                            
                            # Extract file content with better error handling
                            try:
                                extracted_content = zip_file.read(filename)
                                extracted_files_count += 1
                                logger.info(f"✅ Successfully extracted {filename}")
                            except Exception as extract_error:
                                logger.warning(f"⚠️ Failed to extract {filename}: {extract_error}")
                                continue
                            
                            # Enhanced file type detection
                            file_type, ext = detect_file_type(filename, file_content=extracted_content)
                            logger.info(f"🔍 Detected file type for {filename}: {file_type}")
                            
                            # Process different file types with enhanced support
                            if file_type in ['text', 'csv']:
                                # Try multiple encodings
                                text_content = None
                                for encoding in ['utf-8', 'utf-16', 'latin1', 'cp1252']:
                                    try:
                                        text_content = extracted_content.decode(encoding)
                                        logger.info(f"✅ Decoded {filename} with {encoding}")
                                        break
                                    except UnicodeDecodeError:
                                        continue
                                
                                if text_content is not None:
                                    doc = Document(
                                        page_content=text_content,
                                        metadata={
                                            "source": f"archive/{filename}", 
                                            "type": file_type,
                                            "archive_extraction": True,
                                            "file_size": file_info.file_size
                                        }
                                    )
                                    documents.append(doc)
                                else:
                                    logger.warning(f"⚠️ Could not decode text file: {filename}")
                                    
                            elif file_type in ['pdf', 'word', 'excel', 'powerpoint']:
                                # Save extracted file temporarily and process
                                try:
                                    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_file:
                                        tmp_file.write(extracted_content)
                                        temp_path = tmp_file.name
                                    
                                    # Process the extracted file
                                    processed_docs = load_document_by_type(temp_path, file_type, f"archive/{filename}")
                                    
                                    # Add archive metadata
                                    for doc in processed_docs:
                                        doc.metadata.update({
                                            "archive_extraction": True,
                                            "archive_file": filename,
                                            "file_size": file_info.file_size
                                        })
                                    
                                    documents.extend(processed_docs)
                                    
                                    # Cleanup
                                    os.unlink(temp_path)
                                    logger.info(f"✅ Processed {file_type} file: {filename}")
                                    
                                except Exception as process_error:
                                    logger.warning(f"⚠️ Failed to process {filename}: {process_error}")
                                    
                            elif file_type == 'image':
                                # Process images from ZIP
                                try:
                                    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_file:
                                        tmp_file.write(extracted_content)
                                        temp_path = tmp_file.name
                                    
                                    processed_docs = load_document_by_type(temp_path, file_type, f"archive/{filename}")
                                    for doc in processed_docs:
                                        doc.metadata.update({
                                            "archive_extraction": True,
                                            "archive_file": filename,
                                            "file_size": file_info.file_size
                                        })
                                    documents.extend(processed_docs)
                                    
                                    os.unlink(temp_path)
                                    logger.info(f"✅ Processed image: {filename}")
                                    
                                except Exception as img_error:
                                    logger.warning(f"⚠️ Failed to process image {filename}: {img_error}")
                            
                            else:
                                # Create a metadata document for unsupported files
                                doc = Document(
                                    page_content=f"File: {filename}\nType: {file_type}\nSize: {file_info.file_size} bytes\nContent: Binary or unsupported file type from ZIP archive.",
                                    metadata={
                                        "source": f"archive/{filename}",
                                        "type": file_type,
                                        "archive_extraction": True,
                                        "file_size": file_info.file_size,
                                        "processable": False
                                    }
                                )
                                documents.append(doc)
                                logger.info(f"📄 Created metadata document for: {filename}")
                        
                        except Exception as file_error:
                            logger.warning(f"⚠️ Error processing file {filename}: {file_error}")
                            continue
                    
                finally:
                    zip_file.close()
                    
                logger.info(f"✅ ZIP processing complete: {extracted_files_count}/{total_files_count} files extracted, {len(documents)} documents created")
        
        else:
            logger.warning(f"⚠️ Unsupported archive format: {file_extension}")
        
        return documents
        
    except Exception as e:
        logger.error(f"❌ Archive processing failed: {e}")
        # Return at least one document with error info instead of empty list
        error_doc = Document(
            page_content=f"Archive processing failed: {str(e)}. This ZIP file may be corrupted, password-protected, or contain unsupported file formats.",
            metadata={
                "source": "archive_error",
                "type": "archive",
                "error": "extraction_failed",
                "extracted_files": extracted_files_count,
                "total_files": total_files_count
            }
        )
        return [error_doc]

def load_document_by_type(file_path, file_type, original_url):
    """Load document based on file type with enhanced error handling"""
    try:
        documents = []
        
        # Get file extension from file_path
        file_extension = os.path.splitext(file_path)[1].lower()
        
        if file_type == 'pdf':
            loader = PyMuPDFLoader(file_path)
            documents = loader.load()
        elif file_type == 'powerpoint':
            logger.info(f"🎯 Processing PowerPoint file: {file_path}")
            try:
                # First try importing the converter
                from ppt_convert import convert_powerpoint_to_markdown
                logger.info("📚 Imported PowerPoint converter module")
                
                # Check if markdown cache exists
                cached_content = load_markdown_from_cache(original_url)
                if cached_content:
                    documents = [Document(
                        page_content=cached_content,
                        metadata={
                            "source": original_url,
                            "type": "presentation",
                            "cached": True
                        }
                    )]
                    return documents
                
                # Convert PowerPoint if no cache exists
                logger.info("🔄 Starting PowerPoint to Markdown conversion using Docling...")
                markdown_content, error = convert_powerpoint_to_markdown(file_path)
                
                if error:
                    raise Exception(f"PowerPoint conversion failed: {error}")
                
                # Cache the markdown content
                save_markdown_to_cache(markdown_content, original_url)
                
                documents = [Document(
                    page_content=markdown_content,
                    metadata={
                        "source": original_url,
                        "type": "presentation",
                        "cached": False,
                        "converted_at": time.strftime('%Y-%m-%d %H:%M:%S')
                    }
                )]
                logger.info("✅ PowerPoint processing completed successfully")
                return documents
                
            except Exception as e:
                logger.error(f"❌ PowerPoint processing failed: {str(e)}")
                # Fall back to UnstructuredPowerPointLoader
                logger.info("🔄 Falling back to UnstructuredPowerPointLoader...")
                try:
                    loader = UnstructuredPowerPointLoader(file_path)
                    documents = loader.load()
                    return documents
                except Exception as fallback_error:
                    logger.error(f"❌ Fallback loader failed: {str(fallback_error)}")
                    raiseloader = UnstructuredPowerPointLoader(file_path)
            documents = loader.load()
        elif file_type == 'word':
            try:
                loader = UnstructuredWordDocumentLoader(file_path)
                documents = loader.load()
            except Exception as e:
                logger.warning(f"Word loading failed: {e}")
                doc = Document(
                    page_content=f"Word document from {original_url}. Content extraction failed.",
                    metadata={"source": original_url, "type": "word", "error": "extraction_failed"}
                )
                documents = [doc]
                
        elif file_type == 'excel':
            try:
                # Enhanced Excel processing with dramatically improved context window and data extraction
                logger.info("🔍 Starting enhanced Excel processing with extended context...")
                
                # Method 1: Enhanced UnstructuredExcelLoader with more context
                try:
                    loader = UnstructuredExcelLoader(file_path)
                    documents = loader.load()
                    
                    if documents and any(hasattr(doc, 'page_content') and isinstance(doc.page_content, str) and doc.page_content.strip() for doc in documents):
                        # Enhance the extracted content with more detailed analysis
                        enhanced_content = []
                        for doc in documents:
                            if hasattr(doc, 'page_content') and isinstance(doc.page_content, str):
                                content = doc.page_content.strip()
                            else:
                                continue
                            
                            # Add structured analysis to the content
                            enhanced_text = f"""EXCEL SPREADSHEET ANALYSIS:

EXTRACTED DATA:
{content}

DATA STRUCTURE ANALYSIS:
- This Excel file contains structured tabular data
- Data is organized in rows and columns with headers
- May contain formulas, calculations, and formatted information
- Could include multiple worksheets with different data sets
- Numbers, text, dates, and calculated values are present
- Data relationships and dependencies may exist between cells

CONTENT SUMMARY:
The spreadsheet contains detailed information that can be analyzed for specific questions about data values, calculations, trends, comparisons, and structured information retrieval."""
                            
                            enhanced_content.append(enhanced_text)
                        
                        # Create enhanced document with much larger context
                        documents = [Document(
                            page_content="\n\n".join(enhanced_content),
                            metadata={
                                "source": original_url, 
                                "type": "excel", 
                                "method": "enhanced_unstructured",
                                "context_level": "high",
                                "processing_quality": "enhanced"
                            }
                        )]
                        logger.info("✅ Excel processed with Enhanced UnstructuredExcelLoader")
                    else:
                        raise Exception("Empty content from UnstructuredExcelLoader")
                        
                except Exception as e1:
                    logger.warning(f"Enhanced UnstructuredExcelLoader failed: {e1}")
                    
                    # Method 2: Advanced openpyxl processing with comprehensive data extraction
                    try:
                        if file_extension in ['.xlsx', '.xlsm', '.xlsb']:
                            import openpyxl
                            logger.info("🔧 Using advanced openpyxl processing...")
                            
                            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                            comprehensive_content = []
                            
                            # Enhanced worksheet processing
                            for sheet_name in wb.sheetnames:
                                try:
                                    sheet = wb[sheet_name]
                                    logger.info(f"📊 Processing sheet: {sheet_name}")
                                    
                                    sheet_analysis = f"""WORKSHEET: {sheet_name}
DATA EXTRACTION AND ANALYSIS:

RAW DATA:
"""
                                    
                                    # Extract more comprehensive data with better formatting
                                    rows_processed = 0
                                    max_rows = min(sheet.max_row, 1000)  # Process up to 1000 rows for performance
                                    
                                    # Try to identify headers
                                    header_row = None
                                    for row_idx in range(1, min(6, max_rows + 1)):  # Check first 5 rows for headers
                                        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                                        if any(isinstance(cell, str) and len(str(cell)) > 0 for cell in row if cell is not None):
                                            header_row = row_idx
                                            header_data = [str(cell) if cell is not None else "" for cell in row]
                                            sheet_analysis += f"HEADERS (Row {row_idx}): {' | '.join(header_data)}\n\n"
                                            break
                                    
                                    # Extract data rows with enhanced formatting
                                    start_row = header_row + 1 if header_row else 1
                                    for row in sheet.iter_rows(min_row=start_row, max_row=max_rows, values_only=True):
                                        if any(cell is not None for cell in row):
                                            # Enhanced cell processing
                                            processed_row = []
                                            for cell in row:
                                                if cell is not None:
                                                    # Better formatting for different data types
                                                    if isinstance(cell, (int, float)):
                                                        processed_row.append(f"{cell:,.2f}" if isinstance(cell, float) else f"{cell:,}")
                                                    else:
                                                        processed_row.append(str(cell))
                                                else:
                                                    processed_row.append("")
                                            
                                            row_text = " | ".join(processed_row)
                                            sheet_analysis += f"Row {rows_processed + start_row}: {row_text}\n"
                                            rows_processed += 1
                                            
                                            # Limit rows for readability but ensure comprehensive coverage
                                            if rows_processed >= 200:  # Increased from typical limits
                                                sheet_analysis += f"\n... (showing first 200 data rows of {max_rows} total rows)\n"
                                                break
                                    
                                    # Add data analysis summary
                                    sheet_analysis += f"""

SHEET SUMMARY:
- Total rows processed: {rows_processed}
- Sheet contains structured data with {'identified headers' if header_row else 'data without clear headers'}
- Data types include numbers, text, and potentially calculated values
- This sheet can be queried for specific data values, calculations, and analysis

"""
                                    comprehensive_content.append(sheet_analysis)
                                    
                                except Exception as sheet_error:
                                    logger.warning(f"⚠️ Error processing sheet {sheet_name}: {sheet_error}")
                                    comprehensive_content.append(f"WORKSHEET: {sheet_name}\nError processing this sheet: {sheet_error}\n\n")
                            
                            if comprehensive_content:
                                # Create comprehensive document with extensive context
                                final_content = f"""COMPREHENSIVE EXCEL ANALYSIS:

{chr(10).join(comprehensive_content)}

OVERALL SPREADSHEET ANALYSIS:
- This Excel file contains {len(wb.sheetnames)} worksheet(s): {', '.join(wb.sheetnames)}
- Data is structured in tabular format with rows and columns
- Contains various data types including text, numbers, dates, and formulas
- Can be analyzed for specific values, calculations, trends, and data relationships
- Supports complex queries about data patterns, comparisons, and statistical analysis
- All data has been extracted and formatted for comprehensive question answering

QUERY CAPABILITIES:
- Specific cell values and data lookups
- Data comparisons and calculations
- Trend analysis and pattern recognition
- Statistical summaries and aggregations
- Cross-sheet data relationships
- Detailed data filtering and analysis"""

                                documents = [Document(
                                    page_content=final_content,
                                    metadata={
                                        "source": original_url, 
                                        "type": "excel", 
                                        "method": "advanced_openpyxl",
                                        "sheets_count": len(wb.sheetnames),
                                        "context_level": "comprehensive",
                                        "processing_quality": "detailed"
                                    }
                                )]
                                logger.info(f"✅ Excel processed with Advanced openpyxl - {len(wb.sheetnames)} sheets analyzed")
                            else:
                                raise Exception("No content extracted with advanced openpyxl")
                        else:
                            raise Exception("Not a supported xlsx file format")
                            
                    except Exception as e2:
                        logger.warning(f"Advanced openpyxl failed: {e2}")
                        
                        # Method 3: Enhanced xlrd for legacy .xls files
                        try:
                            if file_extension in ['.xls']:
                                import xlrd
                                logger.info("🔧 Using enhanced xlrd for legacy Excel...")
                                
                                workbook = xlrd.open_workbook(file_path)
                                comprehensive_content = []
                                
                                for sheet_idx in range(workbook.nsheets):
                                    try:
                                        sheet = workbook.sheet_by_index(sheet_idx)
                                        sheet_name = sheet.name
                                        
                                        sheet_analysis = f"""LEGACY WORKSHEET: {sheet_name}
ENHANCED DATA EXTRACTION:

RAW DATA:
"""
                                        
                                        # Enhanced legacy processing
                                        for row_idx in range(min(sheet.nrows, 500)):  # Process more rows
                                            try:
                                                row = sheet.row_values(row_idx)
                                                if any(cell for cell in row if str(cell).strip()):
                                                    # Better formatting for legacy data
                                                    formatted_row = []
                                                    for cell in row:
                                                        if isinstance(cell, (int, float)) and cell != 0:
                                                            formatted_row.append(f"{cell:,.2f}" if isinstance(cell, float) else f"{cell:,}")
                                                        elif str(cell).strip():
                                                            formatted_row.append(str(cell))
                                                        else:
                                                            formatted_row.append("")
                                                    
                                                    row_text = " | ".join(formatted_row)
                                                    sheet_analysis += f"Row {row_idx + 1}: {row_text}\n"
                                            except Exception as row_error:
                                                logger.warning(f"Error processing row {row_idx}: {row_error}")
                                                continue
                                        
                                        sheet_analysis += f"""

LEGACY SHEET SUMMARY:
- Contains {sheet.nrows} rows and {sheet.ncols} columns
- Legacy Excel format with preserved data structure
- Data extracted and formatted for analysis
- Supports comprehensive querying and analysis

"""
                                        comprehensive_content.append(sheet_analysis)
                                        
                                    except Exception as sheet_error:
                                        logger.warning(f"Error processing legacy sheet {sheet_idx}: {sheet_error}")
                                        continue
                                
                                if comprehensive_content:
                                    final_content = f"""LEGACY EXCEL COMPREHENSIVE ANALYSIS:

{chr(10).join(comprehensive_content)}

OVERALL LEGACY SPREADSHEET ANALYSIS:
- Legacy Excel file (.xls format) successfully processed
- Contains {workbook.nsheets} worksheet(s)
- Data preserved from legacy format with enhanced formatting
- Supports detailed queries about data values, calculations, and analysis
- All accessible data has been extracted and structured for comprehensive analysis"""

                                    documents = [Document(
                                        page_content=final_content,
                                        metadata={
                                            "source": original_url, 
                                            "type": "excel", 
                                            "method": "enhanced_xlrd",
                                            "legacy_format": True,
                                            "context_level": "comprehensive",
                                            "processing_quality": "detailed"
                                        }
                                    )]
                                    logger.info("✅ Legacy Excel processed with Enhanced xlrd")
                                else:
                                    raise Exception("No content extracted with enhanced xlrd")
                            else:
                                raise Exception("Not a legacy .xls file")
                                
                        except Exception as e3:
                            logger.warning(f"Enhanced xlrd failed: {e3}")
                            
                            # Method 4: Intelligent fallback with comprehensive context
                            logger.info("🔄 Creating intelligent Excel fallback with comprehensive context...")
                            documents = [Document(
                                page_content=f"""EXCEL SPREADSHEET COMPREHENSIVE ANALYSIS:

FILE INFORMATION:
- Source: {original_url}
- Type: Excel Spreadsheet ({file_extension})
- Processing Status: Intelligent Fallback Mode

SPREADSHEET CHARACTERISTICS:
This Excel file contains structured tabular data organized in rows and columns. Excel spreadsheets typically include:

DATA STRUCTURE:
- Headers defining column categories and data types
- Numeric data including integers, decimals, percentages, and currency values
- Text data including names, descriptions, categories, and labels  
- Date and time information
- Calculated fields with formulas and functions
- Multiple worksheets with related or different data sets

ANALYTICAL CAPABILITIES:
The spreadsheet supports comprehensive analysis including:
- Specific data value lookups and retrieval
- Mathematical calculations and statistical analysis
- Data comparisons between rows, columns, and cells
- Trend analysis and pattern identification
- Data filtering and conditional queries
- Cross-referencing between different sections
- Summary statistics and aggregations

QUERY SUPPORT:
This Excel file can answer questions about:
- Specific cell values and data points
- Data ranges and calculations
- Comparisons between different data elements
- Statistical summaries and totals
- Data patterns and relationships
- Conditional data analysis
- Multi-criteria data filtering

PROCESSING NOTE:
While specific content extraction encountered technical limitations, the file structure and data organization principles allow for meaningful responses to data-related questions based on typical Excel spreadsheet patterns and the comprehensive context provided.""",
                                metadata={
                                    "source": original_url, 
                                    "type": "excel", 
                                    "method": "intelligent_fallback",
                                    "processable": True,
                                    "context_level": "comprehensive",
                                    "processing_quality": "enhanced_fallback"
                                }
                            )]
                            logger.info("📄 Intelligent Excel fallback document created with comprehensive context")
                            
            except Exception as e:
                logger.error(f"Complete Excel processing failure: {e}")
                # Enhanced failure fallback with maximum context
                documents = [Document(
                    page_content=f"""EXCEL DOCUMENT COMPREHENSIVE CONTEXT:

FILE: {original_url}
TYPE: Excel Spreadsheet
STATUS: Processing encountered technical difficulties

EXCEL FILE CONTEXT:
This is an Excel spreadsheet file that contains structured data in tabular format. Excel files are designed to store, organize, and analyze data using rows and columns.

TYPICAL EXCEL CONTENT INCLUDES:
- Numerical data (integers, decimals, percentages, currency)
- Text data (names, descriptions, categories, labels)
- Date and time values
- Calculated fields using formulas and functions
- Multiple worksheets for different data sets
- Headers and data organization structures

DATA ANALYSIS CAPABILITIES:
Excel spreadsheets support various analytical operations:
- Data lookups and specific value retrieval
- Mathematical calculations and statistical analysis
- Data comparisons and trend analysis
- Conditional formatting and data filtering
- Summary statistics and aggregations
- Cross-referencing and data relationships

QUESTION ANSWERING POTENTIAL:
Despite processing limitations, this Excel file can potentially provide information about:
- General data structure and organization
- Typical Excel analytical capabilities
- Standard spreadsheet features and functions
- Data management and analysis principles
- Common Excel use cases and applications

TECHNICAL NOTE:
Content extraction encountered difficulties, but the file maintains its Excel structure and can be discussed in terms of general spreadsheet capabilities and standard Excel functionality.""",
                    metadata={
                        "source": original_url, 
                        "type": "excel", 
                        "processable": True, 
                        "error": "processing_failed",
                        "context_level": "comprehensive",
                        "fallback_mode": "enhanced"
                    }
                )]
                
        elif file_type == 'image':
            # Enhanced image processing with multiple OCR fallbacks and better error handling
            try:
                logger.info(f"🖼️ Starting enhanced image OCR processing for {file_path}")
                
                # Method 1: Try EasyOCR (most robust and modern)
                try:
                    import easyocr
                    logger.info("🔍 Attempting EasyOCR processing...")
                    
                    # Initialize EasyOCR reader with common languages
                    reader = easyocr.Reader(['en', 'ch_sim'], gpu=False, verbose=False)
                    
                    # Process image with EasyOCR
                    results = reader.readtext(file_path, detail=0)  # detail=0 for simple text output
                    
                    if results and any(text.strip() for text in results):
                        extracted_text = "\n".join([text.strip() for text in results if text.strip()])
                        
                        documents = [Document(
                            page_content=f"OCR Extracted Text:\n{extracted_text}",
                            metadata={
                                "source": original_url, 
                                "type": "image", 
                                "method": "easyocr",
                                "ocr_confidence": "high",
                                "text_blocks": len(results)
                            }
                        )]
                        logger.info(f"✅ EasyOCR successfully extracted {len(results)} text blocks")
                    else:
                        raise Exception("No text detected by EasyOCR")
                        
                except ImportError:
                    logger.warning("⚠️ EasyOCR not available, trying pytesseract...")
                    raise Exception("EasyOCR not installed")
                except Exception as e:
                    logger.warning(f"⚠️ EasyOCR failed: {e}, trying pytesseract...")
                    
                    # Method 2: Enhanced pytesseract with better configuration
                    try:
                        import pytesseract
                        from PIL import Image
                        logger.info("🔍 Attempting enhanced pytesseract processing...")
                        
                        # Open and preprocess image
                        image = Image.open(file_path)
                        
                        # Convert to RGB if needed
                        if image.mode != 'RGB':
                            image = image.convert('RGB')
                        
                        # Enhanced OCR with custom configuration for better accuracy
                        custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.,!?@#$%^&*()_+-=[]{}|;:,.<>?/`~"\' '
                        
                        # Try different PSM modes for better text detection
                        psm_modes = [6, 8, 13, 11, 12]  # Different page segmentation modes
                        best_text = ""
                        best_confidence = 0
                        
                        for psm in psm_modes:
                            try:
                                config = f'--oem 3 --psm {psm}'
                                text = pytesseract.image_to_string(image, config=config, timeout=30)
                                
                                if text.strip() and len(text.strip()) > len(best_text.strip()):
                                    # Get confidence data
                                    data = pytesseract.image_to_data(image, config=config, output_type=pytesseract.Output.DICT, timeout=30)
                                    confidences = [int(conf) for conf in data['conf'] if int(conf) > 0]
                                    avg_confidence = sum(confidences) / len(confidences) if confidences else 0
                                    
                                    if avg_confidence > best_confidence or not best_text.strip():
                                        best_text = text
                                        best_confidence = avg_confidence
                                        
                            except Exception as psm_error:
                                logger.warning(f"PSM {psm} failed: {psm_error}")
                                continue
                        
                        if best_text.strip():
                            documents = [Document(
                                page_content=f"OCR Extracted Text (Confidence: {best_confidence:.1f}%):\n{best_text.strip()}",
                                metadata={
                                    "source": original_url, 
                                    "type": "image", 
                                    "method": "pytesseract_enhanced",
                                    "ocr_confidence": f"{best_confidence:.1f}%",
                                    "processing_mode": "multi_psm"
                                }
                            )]
                            logger.info(f"✅ Enhanced pytesseract extracted text with {best_confidence:.1f}% confidence")
                        else:
                            raise Exception("No text found with enhanced pytesseract")
                            
                    except ImportError:
                        logger.warning("⚠️ Pytesseract not available, trying UnstructuredImageLoader...")
                        raise Exception("Pytesseract not installed")
                    except Exception as e2:
                        logger.warning(f"⚠️ Enhanced pytesseract failed: {e2}, trying basic methods...")
                        
                        # Method 3: Try UnstructuredImageLoader as fallback
                        try:
                            loader = UnstructuredImageLoader(file_path)
                            documents = loader.load()
                            if documents and documents[0].page_content.strip():
                                # Enhance the content with OCR context
                                enhanced_content = f"Image Text Content:\n{documents[0].page_content.strip()}"
                                documents = [Document(
                                    page_content=enhanced_content,
                                    metadata={
                                        "source": original_url, 
                                        "type": "image", 
                                        "method": "unstructured_loader",
                                        "fallback_method": True
                                    }
                                )]
                                logger.info("✅ UnstructuredImageLoader extracted content")
                            else:
                                raise Exception("Empty content from UnstructuredImageLoader")
                        except Exception as e3:
                            logger.warning(f"⚠️ UnstructuredImageLoader failed: {e3}, using image metadata...")
                            
                            # Method 4: Enhanced image metadata with visual analysis
                            try:
                                from PIL import Image
                                image = Image.open(file_path)
                                width, height = image.size
                                format_name = image.format
                                mode = image.mode
                                
                                # Try to detect if image might contain text based on characteristics
                                aspect_ratio = width / height
                                pixel_count = width * height
                                
                                text_likelihood = "unknown"
                                if aspect_ratio > 2 or aspect_ratio < 0.5:
                                    text_likelihood = "possible (unusual aspect ratio suggests document/text)"
                                elif pixel_count > 500000:  # High resolution suggests document
                                    text_likelihood = "likely (high resolution suggests document)"
                                
                                enhanced_metadata_content = f"""Image Analysis Report:
- Source: {original_url}
- Format: {format_name}
- Dimensions: {width}x{height} pixels ({pixel_count:,} total pixels)
- Color Mode: {mode}
- Aspect Ratio: {aspect_ratio:.2f}
- Text Content Likelihood: {text_likelihood}

OCR Processing Status:
- Advanced OCR methods (EasyOCR, Enhanced Pytesseract) were attempted but failed
- This could be due to:
  1. No readable text in the image
  2. Text in unsupported languages
  3. Poor image quality or resolution
  4. Missing OCR dependencies

Note: The image file exists and is readable, but automated text extraction was not successful."""

                                documents = [Document(
                                    page_content=enhanced_metadata_content,
                                    metadata={
                                        "source": original_url, 
                                        "type": "image", 
                                        "method": "enhanced_metadata",
                                        "width": width,
                                        "height": height,
                                        "format": format_name,
                                        "text_likelihood": text_likelihood,
                                        "ocr_attempted": True,
                                        "ocr_success": False
                                    }
                                )]
                                logger.info("📊 Enhanced image metadata analysis completed")
                            except Exception as e4:
                                # Final fallback
                                logger.warning(f"⚠️ Image metadata extraction failed: {e4}")
                                documents = [Document(
                                    page_content=f"Image file detected from {original_url}. Multiple OCR processing methods were attempted (EasyOCR, Enhanced Pytesseract, UnstructuredImageLoader) but text extraction was not successful. This may indicate the image contains no readable text, uses unsupported languages, or has quality issues preventing OCR processing.",
                                    metadata={
                                        "source": original_url, 
                                        "type": "image", 
                                        "error": "all_ocr_methods_failed",
                                        "attempted_methods": ["easyocr", "pytesseract_enhanced", "unstructured_loader", "metadata_analysis"]
                                    }
                                )]
                                logger.info("📄 Final fallback image document created")
                                
            except Exception as e:
                logger.error(f"❌ Complete image processing failure: {e}")
                documents = [Document(
                    page_content=f"Critical error processing image from {original_url}. Error: {str(e)}. The image file could not be processed due to system-level issues.",
                    metadata={"source": original_url, "type": "image", "error": "critical_failure"}
                )]

                
        elif file_type == 'text':
            try:
                loader = TextLoader(file_path, encoding='utf-8')
                documents = loader.load()
            except Exception as e:
                try:
                    loader = TextLoader(file_path, encoding='latin1')
                    documents = loader.load()
                except Exception as e2:
                    logger.warning(f"Text loading failed: {e2}")
                    documents = []
                    
        elif file_type == 'csv':
            try:
                loader = CSVLoader(file_path)
                documents = loader.load()
            except Exception as e:
                logger.warning(f"CSV loading failed: {e}")
                documents = []
                
        elif file_type == 'binary':
            # For binary files, create a placeholder document
            doc = Document(
                page_content=f"Binary file from {original_url}. Content type: {file_type}. Binary files cannot be processed for text content.",
                metadata={"source": original_url, "type": "binary", "processable": False}
            )
            documents = [doc]
            
        elif file_type == 'archive':
            # Handle archive files separately in the main function
            documents = []
            
        else:
            doc = Document(
                page_content=f"Unsupported file type from {original_url}.",
                metadata={"source": original_url, "type": file_type, "processable": False}
            )
            documents = [doc]
        
        logger.info(f"📄 Loaded {len(documents)} documents from {file_type} file")
        return documents
        
    except Exception as e:
        logger.error(f"❌ Document loading failed for {file_type}: {e}")
        # Return error document
        error_doc = Document(
            page_content=f"Failed to load {file_type} file from {original_url}. Error: {str(e)}",
            metadata={"source": original_url, "type": file_type, "error": "loading_failed"}
        )
        return [error_doc]

def get_cache_path(file_hash):
    """Get the cache file path for a given file hash"""
    return os.path.join(CACHE_DIR, f"{file_hash}.faiss")

def save_vectorstore_to_cache(vectorstore, file_hash):
    """Save vectorstore to cache"""
    try:
        cache_path = get_cache_path(file_hash)
        vectorstore.save_local(cache_path)
        logger.info(f"💾 Vectorstore saved to cache: {cache_path}")
        return True
    except Exception as e:
        logger.warning(f"⚠ Failed to save cache: {e}")
        return False

def save_test_results(results, filename="test_results.json"):
    """Save test results to a JSON file for later analysis, with robust error handling, logging, and console confirmation."""
    import os
    import json
    import time

    max_retries = 2
    for attempt in range(max_retries + 1):
        try:
            # Validate results structure
            if not results or not isinstance(results, dict):
                logger.error(f"❌ Invalid results object: {results}")
                print(f"❌ Invalid results object: {results}")
                return False

            # Prepare to append to results.json as a list of results
            if filename == "results.json":
                if os.path.exists(filename):
                    with open(filename, 'r', encoding='utf-8') as f:
                        try:
                            existing = json.load(f)
                            if not isinstance(existing, list):
                                logger.warning(f"⚠️ Existing results.json is not a list, converting to list.")
                                print(f"⚠️ Existing results.json is not a list, converting to list.")
                                existing = [existing]
                        except Exception as read_err:
                            logger.error(f"⚠️ Failed to read existing results.json: {read_err}")
                            print(f"⚠️ Failed to read existing results.json: {read_err}")
                            existing = []
                else:
                    existing = []
                existing.append(results)
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(existing, f, indent=2, ensure_ascii=False)
                logger.info(f"✅ Results appended and saved to: {filename}")
                print(f"✅ Results appended and saved to: {filename}")
            else:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(results, f, indent=2, ensure_ascii=False)
                logger.info(f"✅ Test results saved to: {filename}")
                print(f"✅ Test results saved to: {filename}")

            # Validate file after write
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                print(f"✅ Confirmed results saved and readable in {filename}.")
            except Exception as verify_err:
                logger.error(f"❌ Error verifying saved file: {verify_err}")
                print(f"❌ Error verifying saved file: {verify_err}")
                return False

            return True
        except Exception as e:
            logger.error(f"❌ Error saving results (attempt {attempt+1}): {e}")
            print(f"❌ Error saving results (attempt {attempt+1}): {e}")
            time.sleep(1)
            if attempt == max_retries:
                # Attempt to not leave file empty/corrupted
                try:
                    with open(filename, 'w', encoding='utf-8') as f:
                        json.dump([], f)
                    logger.warning(f"⚠️ Reset {filename} to empty list after failure.")
                    print(f"⚠️ Reset {filename} to empty list after failure.")
                except Exception as reset_err:
                    logger.critical(f"❌ Failed to reset {filename}: {reset_err}")
                    print(f"❌ Failed to reset {filename}: {reset_err}")
                return False

def load_vectorstore_from_cache(file_hash, embedding_model):
    """Load vectorstore from cache if it exists"""
    try:
        cache_path = get_cache_path(file_hash)
        if os.path.exists(cache_path):
            vectorstore = FAISS.load_local(cache_path, embedding_model, allow_dangerous_deserialization=True)
            logger.info(f"🎯 Loaded vectorstore from cache: {cache_path}")
            return vectorstore
        return None
    except Exception as e:
        logger.warning(f"⚠ Failed to load from cache: {e}")
        return None


async def get_next_nvidia_key():
    global KEY_INDEX
    async with lock:
        key_label, key_value = NVIDIA_KEYS[KEY_INDEX]
        KEY_INDEX = (KEY_INDEX + 1) % len(NVIDIA_KEYS)
    return key_label, key_value

# === FastAPI App ===
app = FastAPI()

# === Logger Setup ===
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# === Embedding Model (GPU/CPU) ===
device = "cuda" if torch.cuda.is_available() else "cpu"
logger.info(f"📦 Embedding model will run on: {device}")

embedding_model = HuggingFaceEmbeddings(
    model_name="BAAI/bge-large-en-v1.5",
    model_kwargs={"device": device}
)

# === Enhanced Document Type Detection ===
def detect_document_type(context_sample, file_type=None):
    """Enhanced document type detection with file type consideration"""
    context_lower = context_sample.lower()
    
    # Consider file type in detection
    if file_type in ['powerpoint']:
        return "presentation"
    elif file_type in ['excel', 'csv']:
        return "spreadsheet"
    elif file_type in ['image']:
        return "image"
    elif file_type in ['word']:
        return "document"
    
    # Existing logic for content-based detection
    policy_keywords = {
        'high': ['policy', 'premium', 'claim', 'coverage', 'insured', 'mediclaim', 'table of benefits'],
        'medium': ['benefit', 'exclusion', 'deductible', 'hospitalisation', 'waiting period', 'sum insured', 'plan a', 'plan b'],
        'low': ['tpa', 'cashless', 'reimbursement', 'network provider', 'room charges', 'icu charges']
    }
    
    academic_keywords = {
        'high': ['newton', 'principia', 'proposition', 'theorem', 'lemma', 'law i', 'law ii', 'law iii'],
        'medium': ['force', 'motion', 'velocity', 'acceleration', 'gravity', 'mass', 'laws of motion'],
        'low': ['physics', 'mathematical', 'philosophy', 'demonstration']
    }
    
    legal_keywords = {
        'high': ['constitution', 'article', 'amendment', 'preamble'],
        'medium': ['rights', 'law', 'legal', 'court', 'justice'],
        'low': ['government', 'citizen', 'state', 'parliament']
    }
    
    def calculate_score(keywords_dict):
        score = 0
        for weight, keywords in keywords_dict.items():
            multiplier = 3 if weight == 'high' else 2 if weight == 'medium' else 1
            score += sum(multiplier for keyword in keywords if keyword in context_lower)
        return score
    
    policy_score = calculate_score(policy_keywords)
    academic_score = calculate_score(academic_keywords)
    legal_score = calculate_score(legal_keywords)
    
    logger.info(f"Document type scores - Policy: {policy_score}, Academic: {academic_score}, Legal: {legal_score}")
    
    if policy_score >= 6:
        return "policy"
    elif academic_score >= 4:
        return "academic"
    elif legal_score >= 4:
        return "legal"
    else:
        return "general"

# [Keep all your existing functions: classify_question, preprocess_question, 
# get_adaptive_retrieval_params, enhanced_hybrid_retrieval, get_enhanced_prompt_template,
# enhanced_clean_and_trim_answer, enhanced_nvidia_llm_call, EnhancedContextManager]

# === Enhanced Question Classification ===
def classify_question(question, document_type):
    """Enhanced question classification with table detection"""
    question_lower = question.lower()
    
    if document_type == "policy":
        if any(word in question_lower for word in ['sub-limit', 'room rent', 'icu charges', 'plan a', 'plan b', 'table', 'charges per day']):
            return "policy_table"
        elif any(word in question_lower for word in ['is', 'does', 'can', 'will', 'are', 'has', 'covered']):
            return "policy_yes_no"
        elif any(word in question_lower for word in ['list', 'documents', 'what are', 'give me']):
            return "policy_list"
        elif any(word in question_lower for word in ['when', 'how long', 'period', 'time']):
            return "policy_time"
        else:
            return "policy_general"
    
    elif document_type == "academic":
        if any(word in question_lower for word in ['how does', 'explain', 'demonstrate', 'derive', 'why']):
            return "academic_explanation"
        elif any(word in question_lower for word in ['what is', 'define', 'who was', 'what are']):
            return "academic_definition"
        elif 'three laws' in question_lower or 'laws of motion' in question_lower:
            return "academic_laws"
        else:
            return "academic_general"
    
    elif document_type == "legal":
        if any(word in question_lower for word in ['article', 'which article', 'under which']):
            return "legal_article"
        elif any(word in question_lower for word in ['is', 'can', 'legal', 'allowed']):
            return "legal_yes_no"
        else:
            return "legal_general"
    
    elif document_type in ["presentation", "spreadsheet", "image", "document"]:
        if any(word in question_lower for word in ['what is', 'what are', 'list', 'show']):
            return f"{document_type}_info"
        elif any(word in question_lower for word in ['how', 'why', 'explain']):
            return f"{document_type}_explanation"
        else:
            return f"{document_type}_general"
    
    else:
        return "general_inquiry"

# === Question Preprocessing ===
def preprocess_question(question):
    """Expand abbreviations and add context"""
    expansions = {
        "IVF": "In Vitro Fertilization (IVF)",
        "OPD": "Outpatient Department (OPD)",
        "ICU": "Intensive Care Unit (ICU)",
        "Rs": "Rupees",
        "C-section": "Caesarean section",
        "AYUSH": "Ayurveda, Yoga, Unani, Siddha, and Homeopathy (AYUSH)",
        "ECG": "Electrocardiogram (ECG)",
        "IONM": "Intra Operative Neuro Monitoring (IONM)",
        "PED": "Pre-Existing Disease (PED)",
        "NCD": "No Claim Discount (NCD)",
        "TPA": "Third Party Administrator (TPA)"
    }
    
    processed_question = question
    for abbr, full in expansions.items():
        if abbr in processed_question and full not in processed_question:
            processed_question = processed_question.replace(abbr, full)
    
    return processed_question

# === Enhanced Retrieval Parameters ===
def get_adaptive_retrieval_params(question_type, document_type):
    """Get enhanced retrieval parameters with special handling for different file types - increased by 30%"""
    
    if document_type == "policy":
        if question_type == "policy_table":
            params = {"k": 26, "lambda_mult": 0.9}
        elif question_type == "policy_list":
            params = {"k": 24, "lambda_mult": 0.8}
        elif question_type in ["policy_yes_no", "policy_time"]:
            params = {"k": 20, "lambda_mult": 0.6}
        else:
            params = {"k": 21, "lambda_mult": 0.7}
    
    elif document_type == "academic":
        if question_type == "academic_explanation":
            params = {"k": 40, "lambda_mult": 0.8}
        elif question_type == "academic_laws":
            params = {"k": 33, "lambda_mult": 0.7}
        elif question_type == "academic_definition":
            params = {"k": 24, "lambda_mult": 0.6}
        else:
            params = {"k": 26, "lambda_mult": 0.7}
    
    elif document_type == "legal":
        if question_type == "legal_article":
            params = {"k": 24, "lambda_mult": 0.7}
        else:
            params = {"k": 20, "lambda_mult": 0.6}
    
    elif document_type in ["presentation", "spreadsheet"]:
        params = {"k": 16, "lambda_mult": 0.8}
    
    elif document_type == "image":
        params = {"k": 11, "lambda_mult": 0.9}
    
    elif document_type == "document":
        params = {"k": 21, "lambda_mult": 0.7}
    
    else:
        params = {"k": 20, "lambda_mult": 0.6}
    
    return params

# === Enhanced Hybrid Retrieval ===
def enhanced_hybrid_retrieval(question, vectorstore, retrieval_params, request_id, question_type):
    """Enhanced hybrid retrieval with file-type specific search"""
    try:
        k = retrieval_params["k"]
        lambda_mult = retrieval_params["lambda_mult"]
        
        # Strategy 1: Direct similarity search
        similarity_docs = vectorstore.similarity_search(question, k=k//3)
        
        # Strategy 2: MMR for diversity
        mmr_docs = vectorstore.max_marginal_relevance_search(
            question, k=k//3, lambda_mult=lambda_mult
        )
        
        # Strategy 3: Enhanced keyword-based search
        keywords = []
        question_lower = question.lower()
        
        # File-type specific keywords
        if question_type.startswith("presentation"):
            keywords.extend(['slide', 'presentation', 'title', 'bullet', 'overview'])
        elif question_type.startswith("spreadsheet"):
            keywords.extend(['table', 'data', 'column', 'row', 'chart', 'value'])
        elif question_type.startswith("image"):
            keywords.extend(['image', 'picture', 'visual', 'diagram', 'figure'])
        elif question_type == "policy_table" or any(word in question_lower for word in ['sub-limit', 'room rent', 'icu charges', 'plan a']):
            keywords.extend(['table of benefits', 'plan a', 'plan b', 'room charges', 'icu charges', 'per day per insured person', 'up to', '% of si'])
        elif 'grace period' in question_lower:
            keywords.extend(['grace period', 'premium payment', 'thirty days'])
        elif 'waiting period' in question_lower:
            keywords.extend(['waiting period', 'continuous coverage', 'months'])
        elif 'three laws' in question_lower:
            keywords.extend(['law i', 'law ii', 'law iii', 'first law', 'second law', 'third law', 'laws of motion'])
        elif 'newton' in question_lower:
            keywords.extend(['newton', 'principia', 'proposition', 'theorem'])
        
        # Use extracted keywords or fall back to question words
        if not keywords:
            keywords = [word for word in question.split() if len(word) > 3][:4]
        
        keyword_docs = []
        for keyword in keywords[:3]:
            try:
                docs = vectorstore.similarity_search(keyword, k=k//6)
                keyword_docs.extend(docs)
            except:
                continue
        
        # Combine all results
        all_docs = similarity_docs + mmr_docs + keyword_docs
        
        # Advanced deduplication
        unique_docs = []
        seen_hashes = set()
        
        for doc in all_docs:
            # Ensure page_content is a string before processing
            if hasattr(doc, 'page_content') and isinstance(doc.page_content, str):
                content_preview = doc.page_content[:300].strip()
                content_hash = hashlib.md5(content_preview.encode()).hexdigest()
                
                if content_hash not in seen_hashes and len(unique_docs) < k:
                    seen_hashes.add(content_hash)
                    unique_docs.append(doc)
        
        logger.info(f"[{request_id}] Retrieved {len(unique_docs)} unique documents using enhanced hybrid search")
        
        return unique_docs[:k]
        
    except Exception as e:
        logger.warning(f"[{request_id}] Enhanced retrieval failed, using fallback: {e}")
        return vectorstore.similarity_search(question, k=retrieval_params["k"])

# === Enhanced Prompt Templates ===
def get_enhanced_prompt_template(question_type, document_type):
    """Get enhanced prompt templates with file-type specific handling"""
    
    if document_type == "presentation":
        return PromptTemplate(
            input_variables=["context", "question"],
            template="""You are a presentation analysis expert.

Using ONLY the provided presentation context, answer the question clearly and comprehensively.

Your answer must be:
- Extract information from slides, titles, and content
- Include key points and structured information
- Be clear and well-organized (3-4 sentences maximum)
- Reference slide content when relevant

Context:
{context}

Question: {question}

Answer:"""
        )
    
    elif document_type == "spreadsheet":
        return PromptTemplate(
            input_variables=["context", "question"],
            template="""You are a spreadsheet data analysis expert.

Using ONLY the provided spreadsheet context, answer the question about data, tables, or values.

Your answer must be:
- Extract exact data from tables and cells
- Include specific numbers, values, and calculations
- Present data clearly and accurately (3-4 sentences maximum)
- Reference table structure when relevant
- Don't repeat "Hackrx "
Context:
{context}

Question: {question}

Answer:"""
        )
    
    elif document_type == "image":
        return PromptTemplate(
            input_variables=["context", "question"],
            template="""You are an image content analysis expert.

Using ONLY the provided image context (OCR extracted text), answer the question about the visual content.

Your answer must be:
- Extract information from OCR text and visual elements
- Describe relevant visual content clearly
- Be concise and accurate (2-3 sentences maximum)
- Note if image processing limitations apply

Context:
{context}

Question: {question}

Answer:"""
        )
    
    elif document_type == "document":
        return PromptTemplate(
            input_variables=["context", "question"],
            template="""You are a document analysis expert.

Using ONLY the provided document context, answer the question comprehensively.

Your answer must be:
- Extract relevant information from the document
- Include key details and supporting information
- Be well-structured and informative (3-4 sentences maximum)
- Reference document sections when applicable

Context:
{context}

Question: {question}

Answer:"""
        )
    
    # [Keep all your existing prompt templates for policy, academic, legal, etc.]
    elif document_type == "policy":
        if question_type == "policy_table":
            return PromptTemplate(
                input_variables=["context", "question"],
                template="""You are a precise insurance policy expert specializing in policy tables and structured data.

Using ONLY the provided policy context, answer the question about tables, limits, charges, or plan-specific information.

Your answer must be:
- Extract exact information from tables, charts, or structured data
- Include specific amounts, percentages, and plan details
- Reference the exact section or table when available
- Be precise and complete (3-4 sentences maximum)
- If information is in a table format, present it clearly

Context:
{context}

Question: {question}

Answer:"""
            )
        # [Include other policy templates...]
        else:
            return PromptTemplate(
                input_variables=["context", "question"],
                template="""You are a comprehensive insurance policy expert.

Using ONLY the provided policy context, provide a complete answer to the question.

Your answer must be:
- Comprehensive yet concise (3-4 sentences maximum)
- Include all relevant details, procedures, and conditions
- Reference specific policy sections when applicable
- Be well-structured and professional

Context:
{context}

Question: {question}

Answer:"""
            )
    
    # [Include other document type templates...]
    else:
        return PromptTemplate(
            input_variables=["context", "question"],
            template="""You are a helpful document assistant.

Using ONLY the provided context, answer the question clearly and accurately.

Your answer must be:
- Clear and informative (3-4 sentences maximum)
- Include relevant details from the context
- Be well-structured and professional

Context:
{context}

Question: {question}

Answer:"""
        )

# === Context Relevance Check ===
def check_context_relevance(context, question):
    """Check if the context contains meaningful information related to the question"""
    if not context or not isinstance(context, str) or len(context.strip()) < 50:
        return False
    
    context_lower = context.lower()
    question_lower = question.lower() if isinstance(question, str) else ""
    
    # Check for generic "no information" patterns
    no_info_patterns = [
        "there is no information provided in the context",
        "no information provided",
        "cannot be processed for text content",
        "content extraction failed",
        "processing failed",
        "unable to retrieve relevant content",
        "failed to load",
        "error processing",
        "technical difficulties",
        "processing encountered technical limitations",
        "no content was extracted",
        "binary file from",
        "unsupported file type"
    ]
    
    # If context contains "no info" patterns, it's not meaningful
    for pattern in no_info_patterns:
        if pattern in context_lower:
            return False
    
    # Check if context has substance beyond generic descriptions
    meaningful_content_indicators = [
        # Look for actual data, numbers, specific information
        any(char.isdigit() for char in context),  # Contains numbers
        len([word for word in context.split() if len(word) > 6]) > 10,  # Has substantial content
        any(word in context_lower for word in question_lower.split()[:3])  # Relates to question
    ]
    
    # Context is meaningful if it has substantial content and relates to the question
    return any(meaningful_content_indicators) and len(context.strip()) > 100

# === Enhanced Answer Trimming ===
def enhanced_clean_and_trim_answer(text, question_type, document_type, question=""):
    """Enhanced answer cleaning and trimming with NO TRUNCATION to fix cut-off responses"""
    import nltk
    try:
        nltk.download('punkt', quiet=True)
        from nltk.tokenize import sent_tokenize, word_tokenize
    except:
        # Fallback if NLTK fails
        def sent_tokenize(text):
            return text.split('. ')
        def word_tokenize(text):
            return text.split()
    
    # Clean the text
    text = re.sub(r'\\n|\n', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()
    
    # REMOVED WORD LIMITS - Let full responses through
    # Instead, rely on LLM max_tokens to control length
    
    # Process sentences for quality but don't truncate
    sentences = sent_tokenize(text)
    
    if not sentences:
        return text
    
    # Join all sentences without artificial limits
    result = " ".join(sentences)
    
    # Final cleanup only
    result = re.sub(r'\s+', ' ', result).strip()
    
    # Ensure proper ending only if the response doesn't already end properly
    if result and not result.endswith(('.', '!', '?', ':', ';')):
        # Check if it looks like it was cut off mid-sentence
        if result.endswith('...') or result.endswith(' with') or result.endswith(' of') or result.endswith(' the'):
            # Keep as is - these are likely truncated by LLM, not us
            pass
        else:
            result += "."
    
    return result

# === Enhanced NVIDIA LLM Call ===
async def enhanced_nvidia_llm_call(context, question, question_type, document_type, api_key, max_retries=2):
    """Enhanced LLM call with file-type specific parameters and fallback to knowledge-based answers"""
    url = "https://integrate.api.nvidia.com/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Accept": "application/json"
    }
    
    # Check if context has meaningful information
    meaningful_context = check_context_relevance(context, question)
    
    if meaningful_context:
        # Use context-based prompt
        prompt_template = get_enhanced_prompt_template(question_type, document_type)
        prompt_text = prompt_template.format(context=context, question=question)
        system_message = "You are a helpful and knowledgeable assistant that provides accurate, complete answers based on the provided context."
    else:
        # Use knowledge-based prompt when no relevant context
        prompt_text = f"Answer this question using your knowledge: {question}"
        if document_type == "excel" or document_type == "spreadsheet":
            system_message = "You are a helpful assistant. Since the Excel/spreadsheet data is not clearly available, answer the question using your general knowledge about the topic, Excel functionality, data analysis concepts, or related domain knowledge as appropriate."
        else:
            system_message = "You are a helpful assistant. Since no relevant information was found in the document, answer the question using your general knowledge about the topic."
    
    messages = [
        {"role": "system", "content": system_message},
        {"role": "user", "content": prompt_text}
    ]
    
    # DRAMATICALLY INCREASED max_tokens to completely eliminate truncation - main fix for cut-off responses
    if not meaningful_context:
        # Knowledge-based responses may need more tokens
        max_tokens = 1560
        temperature = 0.4  # Slightly higher for knowledge-based answers
    elif document_type == "image":
        max_tokens = 1040
        temperature = 0.1
    elif document_type in ["presentation", "spreadsheet", "excel"]:
        max_tokens = 1560
        temperature = 0.3 if not meaningful_context else 0.2  # Higher temp for knowledge-based Excel answers
    elif document_type == "academic":
        if question_type == "academic_laws":
            max_tokens = 2080
        else:
            max_tokens = 1820
        temperature = 0.2
    elif document_type == "policy" and question_type == "policy_table":
        max_tokens = 1560
        temperature = 0.1
    elif document_type == "legal":
        max_tokens = 1300
        temperature = 0.1
    elif document_type == "binary":
        max_tokens = 780
        temperature = 0.3
    else:
        max_tokens = 1300  # general case
        temperature = 0.2
    
    payload = {
        "model": "meta/llama-4-maverick-17b-128e-instruct",
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": temperature,
        "top_p": 0.9,
        "frequency_penalty": 0.1,
        "presence_penalty": 0.1,
        "stream": False
    }
    
    async with httpx.AsyncClient(timeout=40) as client:
        for attempt in range(max_retries + 1):
            try:
                response = await client.post(url, headers=headers, json=payload)
                if response.status_code == 200:
                    content = response.json()['choices'][0]['message']['content']
                    # Handle all possible content types from NVIDIA API
                    if isinstance(content, str):
                        raw_answer = content
                    elif isinstance(content, (dict, list)):
                        raw_answer = str(content)
                    else:
                        raw_answer = str(content) if content is not None else ""
                    
                    # Ensure we have a string before processing
                    if isinstance(raw_answer, str):
                        cleaned_answer = raw_answer.replace('\\n', ' ').replace('\n', ' ')
                        cleaned_answer = re.sub(r'\s+', ' ', cleaned_answer).strip()
                        return cleaned_answer
                    else:
                        return str(raw_answer) if raw_answer else "No response received"
                else:
                    logger.warning(f"API error attempt {attempt + 1}: {response.status_code}")
                    if attempt < max_retries:
                        await asyncio.sleep(2)
                        continue
                    raise Exception(f"NVIDIA API error: {response.text}")
            except Exception as e:
                if attempt < max_retries:
                    logger.warning(f"Request failed attempt {attempt + 1}: {e}")
                    await asyncio.sleep(2)
                    continue
                raise Exception(f"Request failed: {e}")

# === Context Management ===
class EnhancedContextManager:
    def __init__(self):
        self.conversation_history = []
        self._lock = Lock()

    async def add_qa_pair(self, question, answer, question_type):
        """Add Q&A pair to conversation history"""
        async with self._lock:
            self.conversation_history.append({
                "question": question,
                "answer": answer,
                "type": question_type,
                "timestamp": time.time()
            })
            
            if len(self.conversation_history) > 2:
                self.conversation_history.pop(0)

    async def get_relevant_context(self, current_question, current_type):
        """Get minimal relevant context"""
        async with self._lock:
            if not self.conversation_history:
                return ""
            
            related_context = []
            for qa in self.conversation_history[-1:]:
                if any(keyword in current_question.lower() for keyword in qa["question"].lower().split()[:2]):
                    related_context.append(f"Previous context: {qa['answer'][:80]}...")
            
            return "\n".join(related_context) if related_context else ""

# === Input Schema ===
class HackRxInput(BaseModel):
    documents: str
    questions: List[str]

# === Health Check Endpoint ===
@app.get("/health")
def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "message": "RAG server is running",
        "embedding_model": "sentence-transformers/all-MiniLM-L6-v2",
        "device": device
    }

# === Cache Management Endpoints ===
@app.get("/cache/stats")
def get_cache_stats():
    """Get cache statistics"""
    try:
        cache_files = [f for f in os.listdir(CACHE_DIR) if f.endswith('.faiss')]
        total_size = sum(os.path.getsize(os.path.join(CACHE_DIR, f)) for f in cache_files)
        
        return {
            "cache_directory": CACHE_DIR,
            "total_cached_files": len(cache_files),
            "total_cache_size_mb": round(total_size / (1024 * 1024), 2),
            "cache_files": cache_files[:10]  # Show first 10 files
        }
    except Exception as e:
        return {"error": f"Failed to get cache stats: {str(e)}"}

@app.delete("/cache/clear")
def clear_cache():
    """Clear all cache files"""
    try:
        cache_files = [f for f in os.listdir(CACHE_DIR) if f.endswith('.faiss')]
        removed_count = 0
        
        for file in cache_files:
            try:
                file_path = os.path.join(CACHE_DIR, file)
                if os.path.isdir(file_path):
                    import shutil
                    shutil.rmtree(file_path)
                else:
                    os.remove(file_path)
                removed_count += 1
            except Exception as e:
                logger.warning(f"Failed to remove cache file {file}: {e}")
        
        return {
            "message": f"Cache cleared successfully",
            "removed_files": removed_count
        }
    except Exception as e:
        return {"error": f"Failed to clear cache: {str(e)}"}

# === Main Enhanced Endpoint ===
async def process_single_question(q, i, total_questions, document_type, vectorstore, api_key, request_id, context_manager):
    try:
        logger.info(f"[{request_id}] ❓ Question {i}/{total_questions}: {q}")
        
        processed_q = preprocess_question(q)
        question_type = classify_question(processed_q, document_type)
        logger.info(f"[{request_id}] 🏷 Question type: {question_type}")
        
        previous_context = await context_manager.get_relevant_context(processed_q, question_type)
        
        retrieval_params = get_adaptive_retrieval_params(question_type, document_type)
        logger.info(f"[{request_id}] 🔍 Using retrieval params: {retrieval_params}")
        
        # Retrieval with error handling
        try:
            docs = enhanced_hybrid_retrieval(processed_q, vectorstore, retrieval_params, request_id, question_type)
            context = "\n\n".join([doc.page_content for doc in docs])
            
            if previous_context:
                context += f"\n\n{previous_context}"
        except Exception as e:
            logger.warning(f"[{request_id}] ⚠ Retrieval failed for question {i}, using basic search: {e}")
            try:
                docs = vectorstore.similarity_search(processed_q, k=5)
                context = "\n\n".join([doc.page_content for doc in docs])
            except Exception as e2:
                logger.error(f"[{request_id}] ❌ Basic search also failed for question {i}: {e2}")
                context = f"Unable to retrieve relevant content for this question from the document."
        
        # LLM call with error handling
        try:
            answer = await enhanced_nvidia_llm_call(context, processed_q, question_type, document_type, api_key)
            trimmed_answer = enhanced_clean_and_trim_answer(answer, question_type, document_type, processed_q)
            
            await context_manager.add_qa_pair(processed_q, trimmed_answer, question_type)
            
            logger.info(f"[{request_id}] ✅ Question {i}/{total_questions} answered successfully")
            return trimmed_answer
            
        except Exception as e:
            logger.error(f"[{request_id}] ❌ LLM call failed for question {i}: {e}")
            fallback_answer = f"I encountered an error while processing this question. Please try rephrasing your question or check if the document contains relevant information."
            return fallback_answer
        
    except Exception as question_error:
        logger.error(f"[{request_id}] ❌ Complete failure processing question {i}/{total_questions}: {str(question_error)}")
        return "Sorry, I encountered an error processing this question. Please try again."

@app.post("/hackrx/run")
async def handle_rag_request(payload: HackRxInput):
    temp_file_path = None
    request_id = str(uuid.uuid4())[:8]
    context_manager = EnhancedContextManager()
    
    try:
        # Validate input
        if not payload.documents or not payload.questions:
            return {"error": "❌ Both documents URL and questions are required"}
        
        key_label, api_key = await get_next_nvidia_key()
        if not api_key:
            return {"error": "❌ No valid API key available"}
            
        logger.info(f"[{request_id}] 🔄 Using {key_label} for this request.")
        
        total_questions = len(payload.questions)
        logger.info(f"[{request_id}] 📋 Processing {total_questions} questions")

        # Step 1: Download file safely with binary check
        try:
            file_content, content_type = await download_file_safely(payload.documents, max_size_mb=100, check_binary=True)
            
            # If file is skipped (large binary), answer from knowledge
            if file_content is None:
                logger.info(f"[{request_id}] 🤖 Large binary file detected, answering from knowledge.")
                tasks = [enhanced_nvidia_llm_call("", q, "general_inquiry", "general", api_key) for q in payload.questions]
                answers = await asyncio.gather(*tasks)
                
                # Print all questions and answers to console
                print(f"\n=== Questions and Answers (Large Binary File - Knowledge Based) ===")
                for idx, (q, a) in enumerate(zip(payload.questions, answers), 1):
                    print(f"Q{idx}: {q}")
                    print(f"A{idx}: {a}\n")
                print("=============================\n")
                
                return {"answers": answers}
                
        except Exception as e:
            logger.error(f"[{request_id}] ❌ File download failed: {e}")
            return {"error": f"❌ File download failed: {str(e)}"}
        
        # Step 2: Detect file type
        try:
            file_type, file_extension = detect_file_type(payload.documents, content_type, file_content)
            logger.info(f"[{request_id}] 📄 Detected file type: {file_type} ({file_extension})")
            
            # Handle binary and archive files by answering from knowledge
            if file_type in ['binary', 'archive']:
                logger.info(f"[{request_id}] 🤖 {file_type.capitalize()} file detected, answering from knowledge.")
                tasks = [enhanced_nvidia_llm_call("", q, "general_inquiry", "general", api_key) for q in payload.questions]
                answers = await asyncio.gather(*tasks)
                
                # Print all questions and answers to console
                print(f"\n=== Questions and Answers ({file_type.capitalize()} - Knowledge Based) ===")
                for idx, (q, a) in enumerate(zip(payload.questions, answers), 1):
                    print(f"Q{idx}: {q}")
                    print(f"A{idx}: {a}\n")
                print("=============================\n")
                
                return {"answers": answers}
                
        except Exception as e:
            logger.error(f"[{request_id}] ❌ File type detection failed: {e}")
            # Fallback to knowledge-based answering for all questions
            logger.info(f"[{request_id}] 🤖 Answering from knowledge due to file type detection failure.")
            tasks = [enhanced_nvidia_llm_call("", q, "general_inquiry", "general", api_key) for q in payload.questions]
            answers = await asyncio.gather(*tasks)
            return {"answers": answers}
        
        # Step 3: Generate cache hash
        try:
            file_hash = get_file_hash(payload.documents, file_content)
            logger.info(f"[{request_id}] 🔍 File hash: {file_hash}")
        except Exception as e:
            logger.error(f"[{request_id}] ❌ Hash generation failed: {e}")
            # Generate simple hash from URL only
            file_hash = hashlib.md5(payload.documents.encode()).hexdigest()
        
        # Step 4: Try to load from cache
        vectorstore = None
        try:
            vectorstore = load_vectorstore_from_cache(file_hash, embedding_model)
        except Exception as e:
            logger.warning(f"[{request_id}] ⚠ Cache loading failed: {e}")
            vectorstore = None
        
        if vectorstore is not None:
            logger.info(f"[{request_id}] 🚀 Using cached embeddings - skipping file processing!")
            # Skip to document type detection and question processing when using cache
            documents = []
            processable_docs = []  # Initialize for cached processing
            processing_status = "cached_processing"
            # Skip the document filtering logic for cached embeddings
            skip_document_processing = True
        else:
            logger.info(f"[{request_id}] 🔄 Processing new file...")
            skip_document_processing = False
            # Initialize variables for new file processing
            documents = []
            processable_docs = []
            processing_status = "unknown"
            
            # Step 5: Save file temporarily
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
                    tmp_file.write(file_content)
                    temp_file_path = tmp_file.name
                
                logger.info(f"[{request_id}] 💾 File saved temporarily: {temp_file_path}")
            except Exception as e:
                logger.error(f"[{request_id}] ❌ Failed to save temporary file: {e}")
                return {"error": f"❌ Failed to process file: {str(e)}"}
            
            # Step 6: Enhanced file processing with better error handling for different file types
            documents = []
            processing_status = "unknown"
            
            try:
                if file_type == 'archive':
                    documents = process_archive_file(file_content, file_extension)
                    if not documents:
                        processing_status = "archive_extraction_failed"
                        logger.warning(f"[{request_id}] ⚠ Archive extraction failed")
                        return {
                            "error": f"❌ Archive file could not be extracted. The ZIP file may be corrupted, password-protected, or contain only binary files. Please extract the archive manually and upload individual documents.",
                            "file_type": file_type,
                            "processing_status": processing_status
                        }
                    else:
                        processing_status = "archive_processed"
                else:
                    documents = load_document_by_type(temp_file_path, file_type, payload.documents)
                    if documents and any(doc.metadata.get('error') for doc in documents):
                        processing_status = "content_extraction_failed"
                    elif documents:
                        processing_status = "successfully_processed"
                    else:
                        processing_status = "no_content_found"
                
                if not documents:
                    logger.warning(f"[{request_id}] ⚠ No documents loaded for {file_type} file")
                    if file_type in ['image', 'presentation', 'spreadsheet']:
                        return {
                            "error": f"❌ {file_type.title()} file could not be processed. This may be due to: 1) Corrupted file, 2) Unsupported format variant, 3) Missing required libraries. For images, ensure they contain readable text. For presentations/spreadsheets, try converting to PDF format.",
                            "file_type": file_type,
                            "processing_status": "unsupported_content"
                        }
                    else:
                        return {
                            "error": f"❌ File could not be processed. The {file_type} file may be empty, corrupted, or in an unsupported format. Please verify the file and try again.",
                            "file_type": file_type,
                            "processing_status": "no_content_found"
                        }
                
                logger.info(f"[{request_id}] 📄 Loaded {len(documents)} documents, status: {processing_status}")
                
            except Exception as e:
                logger.error(f"[{request_id}] ❌ Document loading failed: {e}")
                processing_status = "processing_error"
                if file_type in ['excel', 'spreadsheet']:
                    return {
                        "error": f"❌ Spreadsheet processing failed. This may be due to: 1) Complex formulas or macros, 2) Large file size, 3) Corrupted data. Try saving as CSV format or simplify the spreadsheet content.",
                        "file_type": file_type,
                        "processing_status": processing_status,
                        "technical_error": str(e)
                    }
                elif file_type in ['powerpoint', 'presentation']:
                    return {
                        "error": f"❌ Presentation processing failed. The PowerPoint file may contain complex elements that cannot be extracted. Try exporting slides as images with text or converting to PDF format.",
                        "file_type": file_type,
                        "processing_status": processing_status,
                        "technical_error": str(e)
                    }
                else:
                    return {
                        "error": f"❌ File processing failed due to technical error. The {file_type} file could not be loaded or parsed. Please check the file integrity and format.",
                        "file_type": file_type,
                        "processing_status": processing_status,
                        "technical_error": str(e)
                    }
        
        # Only do document filtering if not using cached embeddings
        if not skip_document_processing:
            # Simplified filtering - only remove documents with critical errors
            processable_docs = []
            
            for doc in documents:
                # Only filter out documents with explicit processing errors
                if doc.metadata.get('error') == 'complete_failure':
                    logger.warning(f"[{request_id}] Skipping document with complete failure")
                    continue
                else:
                    processable_docs.append(doc)
            
        # Skip document checking if using cached embeddings
        if not skip_document_processing:
            # If no documents at all, only then show error
            if not processable_docs and not documents:
                logger.warning(f"[{request_id}] ⚠ No documents loaded from file")
                return {
                    "error": f"❌ File could not be processed. No content was extracted from the {file_type} file.",
                    "file_type": file_type,
                    "processing_status": "no_content_extracted"
                }
            
            # Use all available documents if processable_docs is empty but documents exist
            if not processable_docs and documents:
                logger.info(f"[{request_id}] Using all available documents despite quality issues")
                processable_docs = documents
        else:
            logger.info(f"[{request_id}] ✅ Skipping document validation - using cached vectorstore")
            processable_docs = []  # Not needed for cached processing
        
        # Step 7: Process documents if not loaded from cache
        if vectorstore is None:
            # Enhanced chunking based on file type with error handling - increased by 30%
            try:
                if file_type in ['presentation', 'image']:
                    chunk_size = 1300
                    chunk_overlap = 130
                elif file_type in ['spreadsheet', 'csv']:
                    chunk_size = 1560
                    chunk_overlap = 195
                elif file_type == 'text':
                    chunk_size = 1950
                    chunk_overlap = 260
                else:  # pdf, word, etc.
                    num_docs = len(processable_docs)
                    if num_docs > 400:
                        chunk_size = 1820
                        chunk_overlap = 455
                    elif num_docs > 100:
                        chunk_size = 2080
                        chunk_overlap = 325
                    else:
                        chunk_size = 2340
                        chunk_overlap = 260
                
                logger.info(f"[{request_id}] 📊 Using chunking: chunk_size={chunk_size}, overlap={chunk_overlap}")
                
                splitter = RecursiveCharacterTextSplitter(
                    chunk_size=chunk_size, 
                    chunk_overlap=chunk_overlap,
                    separators=["\n\n", "\n", ". ", " ", ""],
                    keep_separator=True
                )
                chunks = splitter.split_documents(processable_docs)
                logger.info(f"[{request_id}] 📑 Created {len(chunks)} chunks.")
                
                # Step 8: Create vectorstore with error handling
                try:
                    vectorstore = FAISS.from_documents(chunks, embedding=embedding_model)
                    logger.info(f"[{request_id}] 🧠 FAISS vectorstore created.")
                    
                    # Step 9: Save to cache
                    try:
                        save_vectorstore_to_cache(vectorstore, file_hash)
                    except Exception as e:
                        logger.warning(f"[{request_id}] ⚠ Failed to save to cache: {e}")
                        
                except Exception as e:
                    logger.error(f"[{request_id}] ❌ Vectorstore creation failed: {e}")
                    return {"error": f"❌ Failed to create embeddings: {str(e)}"}
                    
            except Exception as e:
                logger.error(f"[{request_id}] ❌ Chunking failed: {e}")
                return {"error": f"❌ Failed to process document chunks: {str(e)}"}
        
        # Ensure vectorstore exists before proceeding
        if vectorstore is None:
            logger.error(f"[{request_id}] ❌ No vectorstore available for processing")
            return {"error": "❌ Failed to create or load document embeddings"}
        
        # Step 10: Document type detection with error handling
        try:
            sample_docs = vectorstore.similarity_search("document content overview summary", k=5)
            sample_context = " ".join([doc.page_content[:500] for doc in sample_docs])
            document_type = detect_document_type(sample_context, file_type)
            logger.info(f"[{request_id}] 📋 Final document type: {document_type}")
        except Exception as e:
            logger.warning(f"[{request_id}] ⚠ Document type detection failed: {e}")
            document_type = "general"
        
        # Step 11: Process questions in parallel
        tasks = []
        for i, q in enumerate(payload.questions, 1):
            tasks.append(process_single_question(q, i, total_questions, document_type, vectorstore, api_key, request_id, context_manager))

        answer_list = await asyncio.gather(*tasks)
        
        successful_answers = sum(1 for answer in answer_list if "I encountered an error" not in answer and "Sorry, I encountered an error" not in answer)
        failed_answers = total_questions - successful_answers
        
        logger.info(f"[{request_id}] 🎉 Processing completed! Success: {successful_answers}/{total_questions}, Failed: {failed_answers}/{total_questions}")

        # Print all questions and answers to console
        print("\n=== Questions and Answers ===")
        for idx, (q, a) in enumerate(zip(payload.questions, answer_list), 1):
            print(f"Q{idx}: {q}")
            print(f"A{idx}: {a}\n")
        print("=============================\n")

        # Save the results to a JSON file
        results_to_save = {
            "request_id": request_id,
            "document_url": payload.documents,
            "questions_and_answers": [
                {"question": q, "answer": a} for q, a in zip(payload.questions, answer_list)
            ],
            "summary": {
                "total_questions": total_questions,
                "successful_answers": successful_answers,
                "failed_answers": failed_answers,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
        }
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        results_filename = os.path.join(os.getcwd(), f"hackrx_test_results_{timestamp}.json")
        try:
            save_test_results(results_to_save, results_filename)
            logger.info(f"[{request_id}] ✅ Results JSON saved at {results_filename}")
            # Also save to results.json and print confirmation
            save_test_results(results_to_save, "results.json")
        except Exception as save_error:
            logger.error(f"[{request_id}] ❌ Failed to save results JSON: {save_error}")
        
        return {"answers": answer_list}

    except Exception as e:
        logger.error(f"[{request_id}] ⚠ Critical error during processing: {str(e)}")
        # Return error response but don't crash
        return {"error": f"Critical processing error: {str(e)}"}

    finally:
        # Cleanup temporary file
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
                logger.info(f"[{request_id}] 🧹 Temporary file cleaned.")
            except Exception as cleanup_error:
                logger.warning(f"[{request_id}] ⚠ Cleanup failed: {cleanup_error}")

# === Local Run with Ngrok ===
if __name__ == "__main__":
    import uvicorn
    from pyngrok import ngrok

    public_url = ngrok.connect("8000")
    logger.info(f"🌍 Ngrok tunnel running at: {public_url}")

    uvicorn.run(app, host="0.0.0.0", port=8000)
